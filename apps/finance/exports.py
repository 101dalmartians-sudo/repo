from calendar import month_name
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import Budget, Expense, Income, MonthlyFinancialReport


BRAND_FILL = PatternFill(fill_type='solid', fgColor='A30000')
HEADER_FILL = PatternFill(fill_type='solid', fgColor='E2E8F0')


def _autosize_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 4, 40)


def _style_title(sheet, title, subtitle=''):
    sheet.merge_cells('A1:F1')
    sheet['A1'] = 'Aspire Academy'
    sheet['A1'].fill = BRAND_FILL
    sheet['A1'].font = Font(color='FFFFFF', bold=True, size=16)
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells('A2:F2')
    sheet['A2'] = title
    sheet['A2'].font = Font(bold=True, size=13)
    if subtitle:
        sheet.merge_cells('A3:F3')
        sheet['A3'] = subtitle
        sheet['A3'].alignment = Alignment(horizontal='center')


def _style_header_row(row):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL


def _currency(cell):
    cell.number_format = '$#,##0.00'


def export_expenses_workbook(queryset, filename='expenses.xlsx', title='Expense Report'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Expenses'
    _style_title(sheet, title)
    headers = ['Date', 'Category', 'Description', 'Amount', 'Created By', 'Created Date']
    sheet.append([])
    sheet.append(headers)
    _style_header_row(sheet[5])
    total = Decimal('0.00')
    for expense in queryset:
        sheet.append([
            expense.date,
            expense.category.name,
            expense.description,
            expense.amount,
            expense.created_by.username if expense.created_by else '-',
            expense.created_at.replace(tzinfo=None) if expense.created_at else '',
        ])
        _currency(sheet.cell(sheet.max_row, 4))
        total += expense.amount
    sheet.append(['', '', 'Total', total, '', ''])
    _currency(sheet.cell(sheet.max_row, 4))
    _autosize_columns(sheet)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response


def export_budgets_workbook(queryset, filename='budget-report.xlsx', title='Budget Report'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Budgets'
    _style_title(sheet, title)
    headers = ['Category', 'Period', 'Year', 'Month', 'Budget', 'Spent', 'Remaining', 'Utilization %', 'Status']
    sheet.append([])
    sheet.append(headers)
    _style_header_row(sheet[5])
    for budget in queryset:
        sheet.append([
            budget.category.name,
            budget.get_period_type_display(),
            budget.year,
            month_name[budget.month] if budget.month else '-',
            budget.amount,
            budget.expenses_total,
            budget.remaining_budget,
            float(budget.utilization_percentage),
            budget.status,
        ])
        for column_index in [5, 6, 7]:
            _currency(sheet.cell(sheet.max_row, column_index))
    _autosize_columns(sheet)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response


def export_profit_and_loss_workbook(year, filename='profit-loss.xlsx'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Profit and Loss'
    _style_title(sheet, 'Profit & Loss Report', f'Year {year}')
    headers = ['Month', 'Income', 'Expenses', 'Net Profit / Loss']
    sheet.append([])
    sheet.append(headers)
    _style_header_row(sheet[5])
    totals = {'income': Decimal('0.00'), 'expenses': Decimal('0.00')}
    for month in range(1, 13):
        income = Income.objects.filter(date__year=year, date__month=month).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        expenses = Expense.objects.filter(date__year=year, date__month=month).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        totals['income'] += income
        totals['expenses'] += expenses
        sheet.append([month_name[month], income, expenses, income - expenses])
        for column_index in [2, 3, 4]:
            _currency(sheet.cell(sheet.max_row, column_index))
    sheet.append(['Total', totals['income'], totals['expenses'], totals['income'] - totals['expenses']])
    for column_index in [2, 3, 4]:
        _currency(sheet.cell(sheet.max_row, column_index))
    _autosize_columns(sheet)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response


def export_monthly_reports_workbook(queryset, filename='monthly-finance-reports.xlsx'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Monthly Reports'
    _style_title(sheet, 'Monthly Financial Reports')
    headers = ['Year', 'Month', 'Total Income', 'Total Expenses', 'Net Profit / Loss', 'Budget Utilization %']
    sheet.append([])
    sheet.append(headers)
    _style_header_row(sheet[5])
    for report in queryset:
        sheet.append([
            report.year,
            month_name[report.month],
            report.total_income,
            report.total_expenses,
            report.net_profit_loss,
            float(report.budget_utilization_percentage),
        ])
        for column_index in [3, 4, 5]:
            _currency(sheet.cell(sheet.max_row, column_index))
    _autosize_columns(sheet)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response


def export_annual_summary_workbook(year, filename='annual-financial-summary.xlsx'):
    queryset = MonthlyFinancialReport.objects.filter(year=year)
    return export_monthly_reports_workbook(queryset, filename=filename)
